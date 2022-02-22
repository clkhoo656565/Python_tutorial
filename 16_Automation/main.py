#First step is install Python package to work with spreadsheets specifically
    # More practical functions for spreadsheets specifically
    # easier to use

import openpyxl
import os

path = r'C:\Users\ckhoo5\OneDrive - DXC Production\Documents\Python\Tutorial\16_Automation'
inv_file = openpyxl.load_workbook(os.path.join(path, "inventory.xlsx"))
product_list = inv_file["Sheet1"]


products_per_supplier = {} #products_per_supplier["key"] = "value"
total_value_per_supplier = {}
products_under_10_inv = {}
print(product_list.max_row)

#range(start from which row, end reading row) + 1 is because default will not read until the last row
for product_row in range(2, product_list.max_row + 1): 
    supplier_name = product_list.cell(product_row, 4).value #.cell can take 2 parameters {row, column}, in this case we are for loop each row, column 4 "Supplier"
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    #calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        #print("adding a new supplier")
        products_per_supplier[supplier_name] = 1
    

    #calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price
    
    #calculation products inventory less than 10
    if inventory < 10:
        products_under_10_inv[product_num] = inventory
        
    #add value for total inventory price
    inventory_price.value =  inventory * price

#save the inventory into new file
inv_file.save("inventory_with_total_value.xlsx")


print(f"Total products per supplier: {products_per_supplier}")
print(f"Total value of inventory per supplier: {total_value_per_supplier}")
print(f"Here is the product have less than 10 inventory: {products_under_10_inv}")
""" print(type(inventory))
print(type(price)) """

#save the inventory into new file
inv_file.save("inventory_with_total_value.xlsx")